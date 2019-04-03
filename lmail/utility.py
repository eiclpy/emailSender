# -+-coding=utf8-+-
import os
import sys

import chardet
import openpyxl
import xlrd


def get_file(filepath):
    if not os.path.exists(filepath):
        print('文件夹路径不存在', filepath)
        sys.exit(0)
    txts = []
    attachments = []
    for fname in os.listdir(filepath):
        if os.path.isdir(os.path.join(filepath, fname)):
            continue
        if os.path.splitext(fname)[1] == '.txt':
            txts.append(fname)
        else:
            attachments.append(os.path.join(filepath, fname))
    op = 0
    if len(txts) < 1:
        print('无法找到正文文件')
        sys.exit(0)
    if len(txts) > 1:
        for it, txt in enumerate(txts, 1):
            print(it, txt, sep='\t')
        op = int(input('选择正文文件'))
        while op < 1 or op > len(txt):
            op = int(input('选择错误，请重新选择'))
        op -= 1
    txt_path = os.path.join(filepath, txts[op])
    with open(txt_path, 'rb') as fp:
        file_data = fp.read()
        file_encoding = chardet.detect(file_data)
        content = file_data.decode(file_encoding['encoding'])
    return content, attachments


class excel:
    def __init__(self):
        self._addrs = []
        self._last = 0
        self._email = []

    @staticmethod
    def is_email(x):
        if isinstance(x, str):
            slt = x.split('@')
            return len(slt) == 2 and ('.' in slt[1])
        return False

    @property
    def emails(self):
        if self._last != len(self._addrs):
            st = set()
            self._email = []
            for each in self._addrs:
                if each not in st:
                    st.add(each)
                    self._email.append(each)
            self._last = len(self._addrs)
        return self._email[:]

    def handle_xlsx(self, ws):
        idx = None
        for i, cell in enumerate(next(ws.rows), 1):
            if 'address' in cell.value.lower() or '邮箱' in cell.value:
                idx = i
                break
        if not idx:
            return ()
        emails = (each.value for rows in ws.iter_rows(
            min_row=2, min_col=idx, max_col=idx) for each in rows)
        return (each.strip() for each in emails if self.is_email(each))

    def handle_xls(self, ws):
        idx = None
        for i, cell in enumerate(ws.row_values(0)):
            if 'address' in cell.lower() or '邮箱' in cell:
                idx = i
                break
        if not idx:
            return ()
        emails = (each for each in ws.col_values(idx))
        return (each.strip() for each in emails if self.is_email(each))

    def add(self, fname, colnums=None):
        real_cols = []
        addr = []

        if isinstance(colnums, str):
            slt = colnums.split(',')
            for rng in slt:
                if '-' in rng:
                    ft = rng.split('-')
                    if len(ft) != 2:
                        return False
                    f, t = list(map(int, ft))
                    real_cols += list(range(f-1, t))
                else:
                    real_cols.append(int(rng)-1)
        elif isinstance(colnums, (list, tuple)):
            real_cols = [x-1 for x in colnums if isinstance(x, int)]

        if '.xlsx' in fname:
            wb = openpyxl.load_workbook(fname)

            if not real_cols:
                addr += list(self.handle_xlsx(wb.active))
            else:
                sname = wb.sheetnames
                for index in real_cols:
                    try:
                        ws = wb.get_sheet_by_name(sname[index])
                    except:
                        continue
                    addr += list(self.handle_xlsx(ws))

        elif '.xls' in fname:
            wb = xlrd.open_workbook(fname)
            if not real_cols:
                addr += list(self.handle_xlsx(wb.sheet_by_index(0)))
            if not real_cols:
                real_cols = list(range(wb.nsheets))
            for index in real_cols:
                try:
                    ws = wb.sheet_by_index(index)
                except:
                    continue
                addr += list(self.handle_xls(ws))

        self._addrs += addr
        print('成功添加 {} 列: {}'.format(
            os.path.basename(fname), ' '.join(map(lambda x: str(x+1), real_cols)) if real_cols else 'all'))
